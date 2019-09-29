#! python3
# cellConverter.py - This program reverses columns and rows so that col 3, row 5
#   is now col 5, row 3

import openpyxl, sys, os

# Initiate variables
wb = openpyxl.load_workbook(sys.argv[1])    # Get the Excel spreadsheet from the user on program call
sheet = wb.active
invertedSheet = wb.create_sheet(title="Inverted")

sheetData = []

# Loop through the original spreadsheet collecting all the data in a list of lists
for row in range(1, sheet.max_row + 1):
    newRow = []
    for col in range(1, sheet.max_column + 1):
        invertedSheet.cell(col, row).value = sheet.cell(row, col).value
    sheetData.append(newRow)

wb.save(sys.argv[1])
wb.close()


#! python3
# multiplicationTable.py - This program takes a number, N, from command line
#   and creates an N x N multiplication table in an Excel sheet.

import openpyxl, sys
from openpyxl.styles import Font

wb = openpyxl.Workbook()    # Create a new blank spreadsheet
sheet = wb['Sheet']

numIn = int(sys.argv[1])
boldFont = Font(bold=True)

# Set default values for the first row up to numIn (skipping the first cell)
# Also, make the font bold
for col in range(1, numIn+1):
    currCell = sheet.cell(row=1, column=col+1)
    currCell.value = col
    currCell.font = boldFont

# Set default values for the first column up to numIn (skipping the first cell)
# Also, make the font bold
for row in range(1, numIn+1):
    currCell = sheet.cell(row=row+1, column=1)
    currCell.value = row
    currCell.font = boldFont

# Do the maths
for i in range(1, sheet.max_row):
    for j in range(1, sheet.max_row):
        sheet.cell(row=i+1, column=j+1).value = i * j

wb.save('excel/' + str(numIn) + 'by' + str(numIn) + '.xlsx')
wb.close()
